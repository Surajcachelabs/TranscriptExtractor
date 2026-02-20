import argparse
import os
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import requests
from openpyxl import Workbook, load_workbook


DEFAULT_INPUT = Path('Data/sample_input.xlsx')
DEFAULT_BASE_URL = 'http://localhost:3000'


def normalize_header(value: Any) -> str:
    if value is None:
        return ''
    return str(value).strip().lower()


def detect_columns(headers: List[Any]) -> Tuple[int, int]:
    normalized = [normalize_header(h) for h in headers]

    # Name / CSM member column
    name_candidates = {
        'name',
        'csm',
        'csm team member',
        'csm_member',
        'csm team members',
    }

    # Video link column
    link_candidates = {
        'video link',
        'video url',
        'drive link',
        'google drive link',
        'link',
        'url',
    }

    name_idx = next((idx for idx, col in enumerate(normalized) if col in name_candidates), -1)
    link_idx = next((idx for idx, col in enumerate(normalized) if col in link_candidates), -1)

    if name_idx == -1:
        raise ValueError('Could not find Name/CSM column in input sheet headers.')
    if link_idx == -1:
        raise ValueError('Could not find Video Link column in input sheet headers.')

    return name_idx, link_idx


def read_input_rows(input_path: Path) -> List[Tuple[str, str]]:
    if not input_path.exists():
        raise FileNotFoundError(f'Input file not found: {input_path}')

    wb = load_workbook(input_path, data_only=True)
    ws = wb.active

    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return []

    headers = list(all_rows[0])
    name_idx, link_idx = detect_columns(headers)

    records: List[Tuple[str, str]] = []
    for row in all_rows[1:]:
        if row is None:
            continue

        name = '' if name_idx >= len(row) or row[name_idx] is None else str(row[name_idx]).strip()
        link = '' if link_idx >= len(row) or row[link_idx] is None else str(row[link_idx]).strip()

        if not name and not link:
            continue
        if not link:
            continue

        records.append((name, link))

    return records


def parse_cookie_input(cookie_input: str) -> Dict[str, str]:
    cookie_input = cookie_input.strip()
    if not cookie_input:
        return {}

    # Full cookie header format: key1=val1; key2=val2
    if '=' in cookie_input and ';' in cookie_input:
        cookies: Dict[str, str] = {}
        parts = [part.strip() for part in cookie_input.split(';') if part.strip()]
        for part in parts:
            if '=' not in part:
                continue
            key, value = part.split('=', 1)
            cookies[key.strip()] = value.strip()
        return cookies

    # Key=value format
    if '=' in cookie_input:
        key, value = cookie_input.split('=', 1)
        return {key.strip(): value.strip()}

    # Token-only format
    token = cookie_input
    return {
        'next-auth.session-token': token,
        '__Secure-next-auth.session-token': token,
    }


def extract_transcript_text(payload: Dict[str, Any]) -> str:
    formatted = payload.get('formattedTranscript')
    if isinstance(formatted, str) and formatted.strip():
        return formatted.strip()

    transcript = payload.get('transcript')
    if isinstance(transcript, dict):
        text = transcript.get('text')
        if isinstance(text, str) and text.strip():
            return text.strip()

        segments = transcript.get('segments')
        if isinstance(segments, list):
            lines: List[str] = []
            for seg in segments:
                if isinstance(seg, dict):
                    seg_text = seg.get('text')
                    if isinstance(seg_text, str) and seg_text.strip():
                        lines.append(seg_text.strip())
            if lines:
                return '\n'.join(lines)

    return ''


def build_session(base_url: str, cookie_input: str) -> requests.Session:
    session = requests.Session()
    session.headers.update({'Content-Type': 'application/json'})

    cookies = parse_cookie_input(cookie_input)
    for key, value in cookies.items():
        session.cookies.set(key, value)

    # Optional session check (helps fail early when cookie is invalid)
    check = session.get(f'{base_url}/api/auth/session', timeout=20)
    if check.status_code != 200:
        raise RuntimeError(
            f'Auth session check failed with status {check.status_code}. '
            'Make sure your NextAuth session cookie is valid.'
        )

    try:
        data = check.json()
    except Exception:
        data = None

    if not isinstance(data, dict) or not data:
        raise RuntimeError('Session cookie is invalid/expired. Please provide a fresh cookie from a signed-in browser.')

    return session


def transcribe_link(session: requests.Session, base_url: str, drive_url: str, timeout_sec: int) -> Tuple[str, str]:
    resp = session.post(
        f'{base_url}/api/transcribe',
        json={'driveUrl': drive_url},
        timeout=timeout_sec,
    )

    content_type = resp.headers.get('content-type', '')
    if 'application/json' in content_type.lower():
        payload = resp.json()
    else:
        payload = {'error': resp.text}

    if resp.status_code != 200:
        error = payload.get('error') if isinstance(payload, dict) else str(payload)
        return '', f'HTTP {resp.status_code}: {error or "Unknown error"}'

    transcript = extract_transcript_text(payload if isinstance(payload, dict) else {})
    if not transcript:
        return '', 'No transcript text returned.'

    return transcript, ''


def write_output(output_path: Path, rows: List[Tuple[str, str, str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Output'
    ws.append(['CSM Team Member', 'Video Link', 'Transcript'])

    for row in rows:
        ws.append(list(row))

    wb.save(output_path)


def main() -> int:
    parser = argparse.ArgumentParser(description='Batch transcribe links from Excel via existing web app API.')
    parser.add_argument('--input', type=Path, default=DEFAULT_INPUT, help='Path to input Excel file.')
    parser.add_argument('--output', type=Path, default=None, help='Path to output Excel file.')
    parser.add_argument('--base-url', default=DEFAULT_BASE_URL, help='Website base URL (default: http://localhost:3000).')
    parser.add_argument('--cookie', default=os.getenv('NEXTAUTH_COOKIE', ''), help='Session cookie string or token.')
    parser.add_argument('--timeout', type=int, default=600, help='Timeout in seconds per transcription request.')
    parser.add_argument('--sleep', type=float, default=0.5, help='Sleep seconds between requests.')
    args = parser.parse_args()

    if not args.cookie:
        print('ERROR: Missing cookie. Provide --cookie or set NEXTAUTH_COOKIE environment variable.', file=sys.stderr)
        return 2

    base_url = args.base_url.rstrip('/')
    output = args.output or Path('Data') / f"transcripts_output_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    try:
        records = read_input_rows(args.input)
    except Exception as exc:
        print(f'ERROR reading input: {exc}', file=sys.stderr)
        return 2

    if not records:
        print('No valid input rows found.')
        return 0

    try:
        session = build_session(base_url, args.cookie)
    except Exception as exc:
        print(f'ERROR creating authenticated session: {exc}', file=sys.stderr)
        return 2

    output_rows: List[Tuple[str, str, str]] = []

    total = len(records)
    for idx, (member, link) in enumerate(records, start=1):
        print(f'[{idx}/{total}] Processing: {member or "(blank name)"}')
        transcript, err = transcribe_link(session, base_url, link, args.timeout)
        if err:
            transcript = f'ERROR: {err}'
        output_rows.append((member, link, transcript))
        if idx < total and args.sleep > 0:
            time.sleep(args.sleep)

    output.parent.mkdir(parents=True, exist_ok=True)
    write_output(output, output_rows)
    print(f'Done. Output saved to: {output}')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
